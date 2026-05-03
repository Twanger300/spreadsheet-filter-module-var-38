"""Вспомогательные функции для чтения и записи CSV- и XLSX-файлов."""

from __future__ import annotations

import csv
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

Row = dict[str, Any]


def read_csv(path: str | Path) -> list[Row]:
    """Читает CSV-файл как список словарей."""
    with open(path, newline="", encoding="utf-8-sig") as file:
        return list(csv.DictReader(file))


def write_csv(path: str | Path, rows: list[Row]) -> None:
    """Записывает строки в CSV. Если строк нет, создает пустой файл."""
    fieldnames = list(rows[0].keys()) if rows else []
    with open(path, "w", newline="", encoding="utf-8-sig") as file:
        writer = csv.DictWriter(file, fieldnames=fieldnames)
        if fieldnames:
            writer.writeheader()
            writer.writerows(rows)


def read_xlsx(path: str | Path, sheet_name: str | None = None) -> list[Row]:
    """Читает первый лист или выбранный лист из XLSX-файла."""
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[sheet_name] if sheet_name else workbook.active
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    rows: list[Row] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        rows.append({str(header): value for header, value in zip(headers, row, strict=False)})
    return rows


def write_xlsx(path: str | Path, rows: list[Row], sheet_name: str = "Отфильтровано") -> None:
    """Записывает строки на лист XLSX-файла."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    if rows:
        headers = list(rows[0].keys())
        sheet.append(headers)
        for row in rows:
            sheet.append([row.get(header) for header in headers])
    workbook.save(path)
